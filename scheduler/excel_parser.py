from __future__ import annotations

from datetime import datetime
from typing import Any

from openpyxl import load_workbook

from utils.logger import get_logger

logger = get_logger(__name__)

REQUIRED_COLUMNS = ["post_type", "scheduled_time"]
ALLOWED_POST_TYPES: dict[str, list[str]] = {
    "image": ["media_url", "message", "location_id"],
    "reel": ["media_url", "message"],
    "story": ["media_url"],
    "carousel": ["media_url", "message", "location_id"],
    "whatsapp_text": ["recipient_id", "message"],
    "whatsapp_image": ["recipient_id", "media_url", "message"],
    "whatsapp_video": ["recipient_id", "media_url", "message"],
    "facebook_post": ["message", "media_url"],
}


def parse_excel(file_path: str) -> list[dict[str, Any]]:
    logger.info("Parsing Excel file | path=%s", file_path)
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        logger.warning("Excel file is empty | path=%s", file_path)
        return []

    headers = [str(h).lower().strip() if h else "" for h in rows[0]]
    logger.debug("Excel headers: %s", headers)
    for col in REQUIRED_COLUMNS:
        if col not in headers:
            raise ValueError(f"Excel must contain column: {col}. Found: {headers}")

    posts: list[dict[str, Any]] = []
    errors: list[str] = []
    for i, row in enumerate(rows[1:], start=2):
        if all(cell is None for cell in row):
            continue
        entry = {}
        for j, header in enumerate(headers):
            val = row[j] if j < len(row) else None
            entry[header] = val

        post_type = str(entry.get("post_type", "")).strip().lower()
        if not post_type:
            errors.append(f"Row {i}: missing post_type")
            continue
        if post_type not in ALLOWED_POST_TYPES:
            errors.append(f"Row {i}: invalid post_type '{post_type}'")
            continue

        scheduled_time = entry.get("scheduled_time")
        if isinstance(scheduled_time, datetime):
            pass
        elif isinstance(scheduled_time, str):
            scheduled_time = datetime.fromisoformat(scheduled_time)
        else:
            errors.append(f"Row {i}: invalid scheduled_time")
            continue
        # rer
        # if caption == "":
        #     caption = generate_caption(post_type)

        posts.append({
            "post_type": post_type,
            "platform": str(entry.get("platform") or _infer_platform(post_type)).strip(),
            "media_url": str(entry.get("media_url") or "").strip(),
            "message": str(entry.get("message") or "").strip(),
            "topic": str(entry.get("topic") or "").strip(),
            "recipient_id": str(entry.get("recipient_id") or "").strip(),
            "location_id": str(entry.get("location_id") or "").strip(),
            "scheduled_time": scheduled_time,
        })

    wb.close()

    if errors:
        logger.error("Excel parsing errors:\n%s", "\n".join(errors))
        raise ValueError(f"Excel parsing errors:\n" + "\n".join(errors))

    logger.info("Excel parsed successfully | rows=%d posts=%d errors=%d", len(rows) - 1, len(posts), len(errors))
    return posts


def _infer_platform(post_type: str) -> str:
    if post_type.startswith("whatsapp"):
        return "whatsapp"
    if post_type.startswith("facebook"):
        return "facebook"
    return "instagram"
