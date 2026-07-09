from __future__ import annotations

from typing import Any

from mcp.server.fastmcp import FastMCP
from openpyxl import load_workbook

from utils.decorators import _tool
from utils.logger import get_logger

logger = get_logger(__name__)


def register_excel_tools(mcp: FastMCP) -> None:
    tool = _tool(mcp)

    @tool
    def update_excel_cell(file_path: str, row: int, column: str, value: str) -> dict[str, Any]:
        """Update a cell in an Excel file. column is the letter (A, B, C...), row starts at 1."""
        wb = load_workbook(file_path)
        ws = wb.active
        ws[f"{column.upper()}{row}"].value = value
        wb.save(file_path)
        logger.info(f"Updated cell {column.upper()}{row} with value {value}")
        wb.close()
        return {"success": True, "cell": f"{column.upper()}{row}", "value": value}

    @tool
    def read_excel_cell(file_path: str, row: int, column: str) -> dict[str, Any]:
        """Read a cell value from an Excel file."""
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        value = ws[f"{column.upper()}{row}"].value
        wb.close()
        return {"cell": f"{column.upper()}{row}", "value": str(value) if value is not None else None}
